""" Module to help export informations. """

from io import BytesIO

from openpyxl import Workbook

from collectives.utils.misc import deepgetattr


def export_roles(roles):
    """Create an excel with the input role and related user.

    :param roles:
    :type roles: array of :py:class:`collectives.models.role.Role`
    :returns: The excel with all info
    :rtype: :py:class:`io.BytesIO`
    """
    workbook = Workbook()
    worksheet = workbook.active
    fields = {
        "user.license": "Licence",
        "user.first_name": "Prénom",
        "user.last_name": "Nom",
        "user.mail": "Email",
        "user.phone": "Téléphone",
        "activity_type.name": "Activité",
        "name": "Role",
    }
    worksheet.append(list(fields.values()))

    for role in roles:
        worksheet.append([deepgetattr(role, field, "-") for field in fields])

    # set column width
    for i in range(ord("A"), ord("A") + len(fields)):
        worksheet.column_dimensions[chr(i)].width = 25

    out = BytesIO()
    workbook.save(out)
    out.seek(0)

    return out


def export_badges(badges):
    """Create an excel with the input badge and related user.

    :param badges:
    :type badges: array of :py:class:`collectives.models.badge.Badge`
    :returns: The excel with all info
    :rtype: :py:class:`io.BytesIO`
    """
    workbook = Workbook()
    worksheet = workbook.active
    fields = {
        "user.license": "Licence",
        "user.first_name": "Prénom",
        "user.last_name": "Nom",
        "user.mail": "Email",
        "user.phone": "Téléphone",
        "activity_type.name": "Activité",
        "name": "Badge",
        "level": "Niveau",
        "expiration_date": "Date Expiration",
    }
    worksheet.append(list(fields.values()))

    for badge in badges:
        worksheet.append([deepgetattr(badge, field, "-") for field in fields])

    # set column width
    for i in range(ord("A"), ord("A") + len(fields)):
        worksheet.column_dimensions[chr(i)].width = 25

    out = BytesIO()
    workbook.save(out)
    out.seek(0)

    return out


def export_users_registered(event):
    """Create an Excel document with the contact information of registered users at an event.

    :param event:
    :type event: :py:class:`collectives.models.event.event`
    :returns: The excel with all info
    :rtype: :py:class:`io.BytesIO`
    """
    workbook = Workbook()
    worksheet = workbook.active
    fields = [
        "Licence",
        "Nom",
        "Téléphone",
        "Email",
        "En cas d'accident",
    ]
    worksheet.append(fields)

    for reg in event.active_registrations():
        temp = [
            reg.user.license,
            reg.user.full_name(),
            reg.user.phone,
            reg.user.mail,
            f"{ reg.user.emergency_contact_name } ({ reg.user.emergency_contact_phone })",
        ]
        worksheet.append(temp)

    # set column width
    for i in range(ord("A"), ord("A") + len(fields)):
        worksheet.column_dimensions[chr(i)].width = 25

    out = BytesIO()
    workbook.save(out)
    out.seek(0)

    return out
