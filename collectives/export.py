from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from flask import current_app

from .models import Event


class XlsCells:
    ACTIVITIES = 'A8'
    TITLE = 'D8'
    DESCRIPTION = 'A10'

    LEADER_NAME = 'E11'
    LEADER_PHONE = 'E12'
    LEADER_EMAIL = 'E13'

    START = 'E15'
    END = 'E16'

    REG_START_ROW = 28
    REG_ROW_COUNT = 12
    def reg_cell(self, index, field):
        column = ''
        if field == 'index':
            column = 'A'
        elif field == 'license':
            column = 'B'
        elif field == 'name':
            column = 'C'
        elif field == 'email':
            column = 'E'
        elif field == 'phone':
            column = 'D'
        return '{c}{r}'.format(r = self.REG_START_ROW+index, c=column)


def strip_tags(text):
    return text


def to_xlsx(event, cells=XlsCells()):
    wb = load_workbook(filename=current_app.config['XLSX_TEMPLATE'])

    # grab the active worksheet
    ws = wb.active

    # Activity types
    activity_types = [at.name for at in event.activity_types]
    ws[cells.ACTIVITIES] = 'Collective de {}'.format(', '.join(activity_types))

    # Title, Description
    ws[cells.TITLE] = event.title
    ws[cells.DESCRIPTION] = strip_tags(event.description)

    # Leader(s)
    leader_names = [l.full_name() for l in event.leaders]
    leader_phones = [l.phone for l in event.leaders if l.phone]
    leader_emails = [l.mail for l in event.leaders]
    ws[cells.LEADER_NAME] = ', '.join(leader_names)
    ws[cells.LEADER_EMAIL] = ', '.join(leader_emails)
    ws[cells.LEADER_PHONE] = ', '.join(leader_phones)

    # Dates
    ws[cells.START] = 'Départ: {d}/{m}/{y}'.format(
        d=event.start.day, m=event.start.month, y=event.start.year)
    ws[cells.END] = 'Retour: {d}/{m}/{y}'.format(
        d=event.start.day, m=event.start.month, y=event.start.year)

    # Participants
    registrations = event.active_registrations()

    # Default templage has a limited number of existing rows
    # If we have more registrations, insert supplemental rows
    reg_count = len(registrations)
    for i in range(cells.REG_ROW_COUNT, reg_count):
        ws.insert_rows(cells.REG_START_ROW+i)
        ws[cells.reg_cell(i, 'index')] = i+1

    # Insert participants data
    for i, reg in enumerate(registrations):
        ws[cells.reg_cell(i, 'name')] = reg.user.full_name()
        ws[cells.reg_cell(i, 'license')] = reg.user.license
        ws[cells.reg_cell(i, 'email')] = reg.user.mail
        ws[cells.reg_cell(i, 'phone')] = reg.user.phone


    return save_virtual_workbook(wb)
