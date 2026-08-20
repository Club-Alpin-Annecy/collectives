"""Module to test the retex (Ptirex) module."""

# pylint: disable=unused-argument

import json
from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

from collectives.models import ActivityType, Event, Retex, RetexStatus, db
from tests import fixtures, utils


def test_edit_retex_as_leader(leader_client, past_event):
    """Test that a leader can create a retex for an event they led."""
    response = leader_client.get(
        f"/retex/event/{past_event.id}/edit", follow_redirects=True
    )
    assert response.status_code == 200

    data = utils.load_data_from_form(response.text, "retex_form")
    data["status"] = [str(int(RetexStatus.Normal))]
    data["description"] = "Belle sortie, tout s'est bien passé."

    response = leader_client.post(
        f"/retex/event/{past_event.id}/edit", data=data, follow_redirects=True
    )
    assert response.status_code == 200
    # Redirected to the event page: the admin panel "Retex" button must render
    assert "Retex" in response.text

    db.session.refresh(past_event)
    assert past_event.retex is not None
    assert past_event.retex.author_id == leader_client.user.id
    assert past_event.retex.status == RetexStatus.Normal
    assert "Belle sortie" in past_event.retex.rendered_description


def test_edit_retex_wrong_user(user1_client, past_event):
    """Test that a non-leader cannot access the retex edition page."""
    response = user1_client.get(f"/retex/event/{past_event.id}/edit")
    assert response.status_code == 302

    response = user1_client.post(
        f"/retex/event/{past_event.id}/edit",
        data={"status": str(int(RetexStatus.Normal)), "description": "Test"},
    )
    assert response.status_code == 302
    assert past_event.retex is None


def test_edit_retex_non_collective_event(leader_client, youth_event):
    """Test that retex is refused for a non "collective" event, even for its leader."""
    response = leader_client.get(f"/retex/event/{youth_event.id}/edit")
    assert response.status_code == 302
    assert youth_event.retex is None


def test_retex_update_reuses_row(leader_client, past_event):
    """Test that editing twice does not create two Retex rows."""
    data = {"status": str(int(RetexStatus.Normal)), "description": "Premier jet"}
    leader_client.post(
        f"/retex/event/{past_event.id}/edit", data=data, follow_redirects=True
    )

    data = {"status": str(int(RetexStatus.Shortened)), "description": "Version finale"}
    leader_client.post(
        f"/retex/event/{past_event.id}/edit", data=data, follow_redirects=True
    )

    db.session.refresh(past_event)
    assert past_event.retex is not None
    assert past_event.retex.status == RetexStatus.Shortened
    assert past_event.retex.description == "Version finale"


def test_retex_prompt_visible_to_leader_when_missing(
    client, leader_user, user1, past_event
):
    """Test that a leader sees a prompt to write the retex when there is none yet,
    but a regular user does not."""
    fixtures.client.login(client, leader_user)
    response = client.get(f"/collectives/{past_event.id}", follow_redirects=True)
    assert response.status_code == 200
    assert "pas encore été rédigé" in response.text
    assert "Rédiger le retex" in response.text

    fixtures.client.login(client, user1)
    response = client.get(f"/collectives/{past_event.id}", follow_redirects=True)
    assert response.status_code == 200
    assert "pas encore été rédigé" not in response.text


def test_retex_visible_on_event_page(client, leader_user, user1, past_event):
    """Test that the retex is visible on the event page to any logged-in user."""
    fixtures.client.login(client, leader_user)
    data = {
        "status": str(int(RetexStatus.NearMissAccident)),
        "description": "Attention au caillou",
    }
    client.post(f"/retex/event/{past_event.id}/edit", data=data, follow_redirects=True)

    fixtures.client.login(client, user1)
    response = client.get(f"/collectives/{past_event.id}", follow_redirects=True)
    assert response.status_code == 200
    assert "Attention au caillou" in response.text


def test_my_retex_list(leader_client, past_event):
    """Test the 'Mes Retex' page and its backing API."""
    response = leader_client.get("/retex/mine")
    assert response.status_code == 200

    response = leader_client.get("/api/retex/mine")
    assert response.status_code == 200
    payload = json.loads(response.text)
    event_ids = [e["id"] for e in payload["data"]]
    assert past_event.id in event_ids
    matching = next(e for e in payload["data"] if e["id"] == past_event.id)
    assert matching["has_retex"] is False

    data = {"status": str(int(RetexStatus.Normal)), "description": "Ok"}
    leader_client.post(
        f"/retex/event/{past_event.id}/edit", data=data, follow_redirects=True
    )

    response = leader_client.get("/api/retex/mine")
    payload = json.loads(response.text)
    matching = next(e for e in payload["data"] if e["id"] == past_event.id)
    assert matching["has_retex"] is True


def test_supervised_retex_access(client, user1, supervisor_user, past_event):
    """Test access restrictions on the activity-supervisor retex page."""
    fixtures.client.login(client, user1)
    response = client.get("/retex/activity_supervision")
    assert response.status_code == 302

    fixtures.client.login(client, supervisor_user)
    response = client.get("/retex/activity_supervision")
    assert response.status_code == 200

    response = client.get("/api/retex/supervised")
    assert response.status_code == 200
    payload = json.loads(response.text)
    event_ids = [e["id"] for e in payload["data"]]
    assert past_event.id in event_ids


def test_export_retex(client, user1, supervisor_user, past_event):
    """Test the Excel export of supervised retex."""
    fixtures.client.login(client, user1)
    response = client.post("/retex/activity_supervision/export")
    assert response.status_code == 302

    fixtures.client.login(client, supervisor_user)
    response = client.post("/retex/activity_supervision/export")
    assert response.status_code == 200
    assert (
        response.headers["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_pending_retex_count(leader_client, past_event):
    """Test User.pending_retex_count()."""
    leader = leader_client.user
    assert leader.pending_retex_count() == 1

    data = {"status": str(int(RetexStatus.Normal)), "description": "Ok"}
    leader_client.post(
        f"/retex/event/{past_event.id}/edit", data=data, follow_redirects=True
    )

    db.session.refresh(leader)
    assert leader.pending_retex_count() == 0


def test_export_retex_date_filter(client, supervisor_user, past_event):
    """Test that the export date filter narrows down included events."""
    old_event = Event()
    old_event.title = "Sortie ancienne"
    old_event.start = date.today() - timedelta(days=100)
    old_event.end = old_event.start
    old_event.num_online_slots = 1
    old_event.event_type = past_event.event_type
    old_event.activity_types = past_event.activity_types
    old_event.leaders = past_event.leaders
    db.session.add(old_event)
    db.session.commit()

    fixtures.client.login(client, supervisor_user)
    data = {
        "date_from": (date.today() - timedelta(days=20)).isoformat(),
        "date_to": date.today().isoformat(),
    }
    response = client.post("/retex/activity_supervision/export", data=data)
    assert response.status_code == 200

    workbook = load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    # Header row + only the recent event; the old one is filtered out
    assert worksheet.max_row == 2
    assert worksheet["B2"].value == past_event.title
    assert worksheet["E1"].value == "Statut"
    assert "Retex rédigé" not in [c.value for c in worksheet[1]]
    assert worksheet["E2"].value == "Retex Absent"


def test_export_retex_activity_filter_rejects_unsupervised_activity(
    client, supervisor_user, past_event
):
    """A supervisor cannot use activity_ids to pull events from activities they do
    not supervise: the form itself must reject the unknown choice."""
    other_activity = ActivityType(name="Ski", short="ski_retex_test", trigram="SKR")
    db.session.add(other_activity)
    db.session.commit()

    fixtures.client.login(client, supervisor_user)
    data = {"activity_ids": [str(other_activity.id)]}
    response = client.post("/retex/activity_supervision/export", data=data)
    assert response.status_code == 302


def test_export_retex_leader_filter(client, supervisor_user, leader2_user, past_event):
    """Test that the leader filter narrows results to events led by that leader."""
    other_event = Event()
    other_event.title = "Sortie d'un autre encadrant"
    other_event.start = past_event.start
    other_event.end = past_event.end
    other_event.num_online_slots = 1
    other_event.event_type = past_event.event_type
    other_event.activity_types = past_event.activity_types
    other_event.leaders = [leader2_user]
    db.session.add(other_event)
    db.session.commit()

    fixtures.client.login(client, supervisor_user)
    leader_id = past_event.leaders[0].id
    response = client.post(
        "/retex/activity_supervision/export", data={"leader_id": str(leader_id)}
    )
    assert response.status_code == 200

    workbook = load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 2
    assert worksheet["B2"].value == past_event.title


def test_export_retex_status_filter(client, supervisor_user, leader2_user, past_event):
    """Test that the status filter narrows results to retex with that status."""
    past_event.retex = Retex(
        author_id=past_event.leaders[0].id,
        status=RetexStatus.Accident,
        description="Chute",
    )
    db.session.add(past_event)

    other_event = Event()
    db.session.add(other_event)
    other_event.title = "Sortie normale"
    other_event.start = past_event.start
    other_event.end = past_event.end
    other_event.num_online_slots = 1
    other_event.event_type = past_event.event_type
    other_event.activity_types = past_event.activity_types
    other_event.leaders = [leader2_user]
    other_event.retex = Retex(
        author_id=leader2_user.id, status=RetexStatus.Normal, description="RAS"
    )
    db.session.commit()

    fixtures.client.login(client, supervisor_user)
    data = {"status_ids": [str(int(RetexStatus.Accident))]}
    response = client.post("/retex/activity_supervision/export", data=data)
    assert response.status_code == 200

    workbook = load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 2
    assert worksheet["B2"].value == past_event.title
