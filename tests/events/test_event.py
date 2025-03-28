"""Event actions tests."""

from datetime import datetime, timedelta
import re
from io import BytesIO

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from flask import url_for

from collectives.models import db, EventStatus, ActivityType, Event, EventVisibility
from collectives.models import RoleIds, Question, QuestionType
from tests import utils
from tests.fixtures.user import promote_user


def test_event_access(user1_client, event):
    """Test regular acces to event description"""
    response = user1_client.get(f"/collectives/{event.id}")
    assert response.status_code == 302
    assert response.headers["Location"] == "/collectives/1-new-collective"

    response = user1_client.get(response.headers["Location"])
    assert response.status_code == 200


def test_activity_event_access(user1_client, activity_event):
    """Test regular acces to event description"""

    # Denied access to activity event
    response = user1_client.get(f"/collectives/{activity_event.id}")
    assert response.status_code == 302
    assert response.headers["Location"] == "/collectives/"

    # Remove activity_event activity_type -- should still not have access
    activity_event.activity_types.clear()
    db.session.commit()

    response = user1_client.get(f"/collectives/{activity_event.id}")
    assert response.status_code == 302
    assert response.headers["Location"] == "/collectives/"

    # Give role to user for this activty -- should get access
    promote_user(user1_client.user, RoleIds.Trainee, "Alpinisme")
    db.session.commit()
    response = user1_client.get(f"/collectives/{activity_event.id}")
    assert response.status_code == 302
    assert response.headers["Location"].startswith(f"/collectives/{activity_event.id}")

    # Remove event's activity -- should keep access
    activity_event.activity_types.clear()
    db.session.commit()
    response = user1_client.get(f"/collectives/{activity_event.id}")
    assert response.headers["Location"].startswith(f"/collectives/{activity_event.id}")

    # Set activity_type to another activity -- should lose access
    not_alpinisme = ActivityType.query.filter(ActivityType.name != "Alpinisme").first()
    activity_event.activity_types.append(not_alpinisme)
    db.session.commit()
    response = user1_client.get(f"/collectives/{activity_event.id}")
    assert response.headers["Location"] == "/collectives/"


def test_leader_event_access(leader_client, event, activity_event):
    """Test leader acces to event description"""
    response = leader_client.get(f"/collectives/{event.id}")
    assert response.status_code == 302
    assert response.headers["Location"] == "/collectives/1-new-collective"

    response = leader_client.get(response.headers["Location"])
    assert response.status_code == 200

    # Ok access to activity event
    response = leader_client.get(f"/collectives/{activity_event.id}")
    assert response.status_code == 302
    assert response.headers["Location"].startswith(f"/collectives/{activity_event.id}")

    # Set activity_type to another activity -- should keep access because leading it
    not_alpinisme = ActivityType.query.filter(ActivityType.name != "Alpinisme").first()
    activity_event.activity_types.clear()
    activity_event.activity_types.append(not_alpinisme)
    db.session.commit()
    response = leader_client.get(f"/collectives/{activity_event.id}")
    assert response.headers["Location"].startswith(f"/collectives/{activity_event.id}")


def test_event_deletion(leader_client, event, event1_with_reg):
    """Test leader delete rights"""

    # Leader can delete their event if no reg
    response = leader_client.post(
        url_for("event.delete_event", event_id=event.id), follow_redirects=True
    )
    assert response.status_code == 200
    assert "error message" not in response.text

    # Leader cannot delete event with reg
    response = leader_client.post(
        url_for("event.delete_event", event_id=event1_with_reg.id),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert "error message" in response.text


def test_supervisor_event_deletion(supervisor_client, event1_with_reg):
    """Test supervisor delete rights"""

    # Supervisor can delete events with registrations
    response = supervisor_client.post(
        url_for("event.delete_event", event_id=event1_with_reg.id),
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert "error message" not in response.text


def test_unauthenticated(client, event1_with_reg):
    """Test acces to event description by an unauthenticated client"""
    response = client.get(f"/collectives/{event1_with_reg.id}")
    assert response.status_code == 302
    assert (
        response.headers["Location"]
        == f"/auth/login?next=%2Fcollectives%2F{event1_with_reg.id}"
    )

    # External event should be visible to unauthenticated users
    event1_with_reg.visibility = EventVisibility.External
    db.session.add(event1_with_reg)
    db.session.commit()

    response = client.get(f"/collectives/{event1_with_reg.id}-some-valid-slug")
    assert response.status_code == 200


def test_crawler(client, event):
    """Test acces to event description by an unauthenticated crawler"""
    headers = {
        "User-Agent": "facebookexternalhit/1.0 (+http://www.facebook.com/externalhit_uatext.php)"
    }

    event.start = datetime(2072, 10, 20, 0, 0, 0)
    event.end = datetime(2072, 10, 20, 0, 0, 0)

    response = client.get(f"/collectives/{event.id}", headers=headers)
    assert response.status_code == 302
    assert response.headers["Location"] == f"/collectives/{event.id}/preview"

    response_preview = client.get(response.headers["Location"], headers=headers)
    assert response_preview.status_code == 200
    soup = BeautifulSoup(response_preview.text, features="lxml")
    assert (
        soup.select_one('meta[property="og:title"]')["content"]
        == "Collectives: New collective "
    )
    assert (
        soup.select_one('meta[property="og:url"]')["content"]
        == f"/collectives/{event.id}-new-collective"
    )
    assert (
        soup.select_one('meta[property="og:image"]')["content"]
        == "http://localhost/static/caf/logo.svg"
    )

    description = soup.select_one('meta[property="og:description"]')["content"]

    assert (
        description == "Collective Alpinisme - jeudi 20 octobre 2072 - 10 places - "
        "Par Romeo CAPO - Lorem ipsum dolor sit amet, consectetur adipiscing elit. "
        "Quisque mollis vitae diam at hendrerit...."
    )

    # Test an event with a photo
    event.photo = "test.jpg"
    response = client.get(response.headers["Location"], headers=headers)
    assert response.status_code == 200
    soup = BeautifulSoup(response.text, features="lxml")
    assert (
        "imgsizer/test.jpg" in soup.select_one('meta[property="og:image"]')["content"]
    )

    # No user agent means not a crawler
    headers = {"User-Agent": ""}
    response = client.get(f"/collectives/{event.id}", headers=headers)
    assert response.status_code == 302
    assert "preview" not in response.headers["Location"]


def test_event_creation(leader_client):
    """Test regular event creation by a leader"""
    response = leader_client.get("/collectives/add")
    assert response.status_code == 200

    now = datetime.now()
    alpinisme = ActivityType.query.filter_by(name="Alpinisme").first()
    data = {
        "update_activity": "0",
        "event_type_id": "1",
        "single_activity_type": alpinisme.id,
        "leader_actions-0-leader_id": leader_client.user.id,
        "main_leader_id": leader_client.user.id,
        "add_leader": "0",
        "update_leaders": "0",
        "title": "Test collectives",
        "status": int(EventStatus.Confirmed),
        "num_slots": "5",
        "start": (now + timedelta(days=10)).strftime("%Y-%m-%d %X"),
        "end": (now + timedelta(days=10, hours=3)).strftime("%Y-%m-%d %X"),
        "num_online_slots": "4",
        "num_waiting_list": "5",
        "registration_open_time": (now - timedelta(days=5)).strftime("%Y-%m-%d %X"),
        "registration_close_time": (now + timedelta(days=5)).strftime("%Y-%m-%d %X"),
        "description": """**Lorem ipsum** dolor sit amet, consectetur adipiscing elit.
                                Cras sodales ut ipsum sit amet ultrices. Curabitur aliquet 
                                id dolor et maximus. Praesent iaculis pretium orci vitae 
                                laoreet. Phasellus laoreet iaculis justo, in vulputate augue
                                bibendum et. _Vestibulum dapibus_ posuere sagittis. Praesent
                                commodo facilisis orci. Sed a volutpat ex. Donec in quam ornare,
                                tempus orci in, semper nisl.""",
        "tag_list": "6",
        "search_terms": "",
        "duplicate_event": "",
        "parent_event_id": "",
        "edit_session_id": "ef32c979-57f6-48f8-a8d5-753050ff2f54",
    }
    response = leader_client.post("/collectives/add", data=data, follow_redirects=True)
    assert response.status_code == 200
    assert "collectives/add" not in response.request.path
    assert data["title"] in response.text
    assert f"Inscrits 0 / {data['num_slots']}" in response.text
    assert "Alpinisme" in response.text
    assert "Handicaf" in response.text
    assert "<strong>Lorem ipsum</strong>" in response.text


def test_event_modification(event, leader_client):
    """Test various event modifications."""
    response = leader_client.get(f"/collectives/{event.id}/edit")
    assert response.status_code == 200

    data = utils.load_data_from_form(response.text, "form_edit_event")
    data["status"] = int(EventStatus.Cancelled)
    response = leader_client.post(
        f"/collectives/{event.id}/edit", data=data, follow_redirects=True
    )
    assert response.status_code == 200
    assert (
        f"collectives/{event.id}-" in response.request.path
    ), "There is an error in request"
    assert "Annulée" in response.text

    data["status"] = int(EventStatus.Pending)
    response = leader_client.post(
        f"/collectives/{event.id}/edit", data=data, follow_redirects=True
    )
    assert response.status_code == 200
    assert (
        f"collectives/{event.id}-" in response.request.path
    ), "There is an error in request"
    assert "En attente" in response.text

    data["description"] = "New **description** for you"
    response = leader_client.post(
        f"/collectives/{event.id}/edit", data=data, follow_redirects=True
    )
    assert response.status_code == 200
    assert (
        f"collectives/{event.id}-" in response.request.path
    ), "There is an error in request"
    assert "New <strong>description</strong> for you" in response.text


def test_event_duplication(leader_client, paying_event):
    """Test the event duplication functionnality"""

    # Add a question to check that it will be correctly duplicated
    paying_event.questions.append(
        Question(
            title="Question",
            description="",
            choices="A\nB\nC\n",
            question_type=QuestionType.MultipleChoices,
            enabled=True,
            required=True,
        )
    )
    db.session.commit()

    response = leader_client.get(f"/collectives/{paying_event.id}/duplicate")
    assert response.status_code == 200

    data = utils.load_data_from_form(response.text, "form_edit_event")
    now = datetime.now()
    data["start"] = (now + timedelta(days=33)).strftime("%Y-%m-%d %X")
    data["end"] = (now + timedelta(days=33, hours=3)).strftime("%Y-%m-%d %X")
    response = leader_client.post("/collectives/add", data=data)
    assert response.status_code == 302
    new_id = re.search(r"/collectives/([0-9]+)", response.location)[1]
    new_event = db.session.get(Event, new_id)

    attributes = [
        "title",
        "rendered_description",
        "photo",
        "num_slots",
        "num_online_slots",
        "registration_open_time",
        "registration_close_time",
        "status",
        "leaders",
        "main_leader",
        "event_type",
        "activity_types",
        "tags",
        "user_group",
    ]
    for attribute in attributes:
        assert getattr(new_event, attribute) == getattr(paying_event, attribute)

    assert paying_event.description in new_event.description

    assert len(new_event.payment_items) == len(paying_event.payment_items)
    assert (
        new_event.payment_items[0].prices[0].amount
        == paying_event.payment_items[0].prices[0].amount
    )
    assert len(new_event.questions) == len(paying_event.questions)
    assert (
        new_event.questions[0].choices_array()
        == paying_event.questions[0].choices_array()
    )


def test_event_export_list(leader_client, event1_with_reg, user2):
    """Test result of user export function"""

    url = f"/collectives/{event1_with_reg.id}/export_registered_users"
    response = leader_client.get(url)
    assert response.status_code == 200
    assert response.content_length > 4000
    assert (
        response.content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 5
    assert worksheet.max_column == 6
    assert worksheet["A3"].value == user2.license
    assert worksheet["B3"].value == user2.first_name
    assert worksheet["C3"].value == user2.last_name.upper()
    assert worksheet["D3"].value == user2.phone
    assert worksheet["E3"].value == user2.mail


def test_event_print(leader_client, event1_with_reg, user2):
    """Test result of user list print function"""
    url = f"/collectives/{event1_with_reg.id}/print"
    response = leader_client.get(url)
    assert response.status_code == 200
    assert event1_with_reg.title in response.text
    assert user2.full_name() in response.text
    assert user2.license in response.text
    assert event1_with_reg.leaders[0].full_name() in response.text
    assert event1_with_reg.rendered_description in response.text
