"""Module to test statistcs engine"""

# pylint: disable=unused-argument

from datetime import timedelta

import openpyxl

from collectives.models import ActivityType, EventType
from collectives.utils.stats import StatisticsEngine
from collectives.utils.time import current_time


def test_statistics_engine_all(stats_env, leader2_user):
    """Tests statistics engine without filter."""
    engine = StatisticsEngine()
    assert engine.nb_registrations() == 8
    assert engine.nb_events() == 6
    assert engine.mean_registrations_per_event() == 8 / 6
    assert engine.nb_events_by_event_type()["Soirée"] == 1
    assert engine.nb_events_by_activity_type()["Canyon"] == 2
    assert engine.nb_events_by_activity_type()["Alpinisme"] == 4
    assert engine.nb_collectives_by_activity_type()["Canyon"] == 1
    assert engine.nb_collectives_by_activity_type()["Alpinisme"] == 4
    assert len(engine.nb_events_by_leaders()) == 2
    assert engine.nb_events_by_leaders()[leader2_user.full_name()] == 1
    assert engine.nb_registrations_by_gender() == {"Femme": 2, "Homme": 6}
    assert engine.nb_events_by_event_tag() == {
        "Handicaf": 3,
        "Rando Cool": 1,
        "Achat": 1,
    }
    assert engine.population_registration_number() == {1: 3, 2: 1, 3: 1}
    assert engine.volunteer_time() == 7
    assert engine.volunteer_time_by_activity_type()["Alpinisme"] == 4
    assert engine.volunteer_time_by_activity_type()["Canyon"] == 3
    assert engine.nb_user_per_activity_type()["Alpinisme"] == 4
    assert engine.nb_active_registrations_per_activity_type()["Alpinisme"] == 6

    assert engine.attendee_time_by_gender_and_license_type() == {
        ("XX", "Femme"): 2,
        ("XX", "Homme"): 5,
        ("E2", "Homme"): 1,
    }

    assert engine.nb_days() is None
    assert engine.mean_events_per_day() is None
    assert engine.mean_registrations_per_day() is None

    assert openpyxl.load_workbook(filename=engine.export_excel())

    engine.nb_user_per_activity_type()


def test_statistics_engine_from_now(stats_env, leader2_user):
    """Tests statistics engine on future events."""
    engine = StatisticsEngine(
        start=current_time(), end=current_time() + timedelta(days=365)
    )
    assert engine.nb_events() == 5
    assert engine.nb_registrations() == 6
    assert engine.mean_registrations_per_event() == 6 / 5
    assert engine.nb_events_by_event_type()["Collective"] == 4
    assert engine.nb_events_by_activity_type()["Canyon"] == 2
    assert engine.nb_events_by_activity_type()["Alpinisme"] == 3
    assert engine.nb_collectives_by_activity_type()["Canyon"] == 1
    assert engine.nb_collectives_by_activity_type()["Alpinisme"] == 3
    assert len(engine.nb_events_by_leaders()) == 2
    assert engine.nb_events_by_leaders()[leader2_user.full_name()] == 1
    assert engine.nb_registrations_by_gender() == {"Femme": 2, "Homme": 4}
    assert engine.nb_events_by_event_tag() == {
        "Handicaf": 2,
        "Rando Cool": 1,
        "Achat": 1,
    }
    assert engine.population_registration_number() == {1: 4, 2: 1}
    assert engine.volunteer_time() == 6
    assert engine.volunteer_time_by_activity_type()["Alpinisme"] == 3
    assert engine.volunteer_time_by_activity_type()["Canyon"] == 3
    assert engine.nb_user_per_activity_type()["Alpinisme"] == 4
    assert engine.attendee_time_by_gender_and_license_type() == {
        ("XX", "Femme"): 2,
        ("XX", "Homme"): 3,
        ("E2", "Homme"): 1,
    }
    assert engine.mean_events_per_day() == 0.0136986301369863
    assert engine.mean_registrations_per_day() == 0.01643835616438356

    assert openpyxl.load_workbook(filename=engine.export_excel())


def test_statistics_engine_only_alpi(stats_env, leader2_user):
    """Tests statistics engine on alpinism events."""
    alpinisme = ActivityType.query.filter_by(name="Alpinisme").first()

    engine = StatisticsEngine(activity_id=alpinisme.id)
    assert engine.nb_events() == 4
    assert engine.nb_registrations() == 6
    assert engine.mean_registrations_per_event() == 6 / 4
    assert engine.nb_events_by_event_type()["Collective"] == 4
    assert "Canyon" not in engine.nb_events_by_activity_type()
    assert engine.nb_events_by_activity_type()["Alpinisme"] == 4
    assert "Canyon" not in engine.nb_collectives_by_activity_type()
    assert engine.nb_collectives_by_activity_type()["Alpinisme"] == 4
    assert len(engine.nb_events_by_leaders()) == 1
    assert leader2_user.full_name() not in engine.nb_events_by_leaders()
    assert engine.nb_registrations_by_gender() == {"Femme": 2, "Homme": 4}
    assert engine.nb_events_by_event_tag() == {"Handicaf": 2}
    assert engine.population_registration_number() == {1: 2, 2: 2}
    assert engine.volunteer_time() == 4
    assert engine.volunteer_time_by_activity_type()["Alpinisme"] == 4
    assert "Canyon" not in engine.volunteer_time_by_activity_type()
    assert engine.attendee_time_by_gender_and_license_type() == {
        ("XX", "Femme"): 2,
        ("XX", "Homme"): 4,
    }
    assert engine.nb_user_per_activity_type()["Alpinisme"] == 4
    assert engine.nb_days() is None
    assert engine.mean_events_per_day() is None
    assert engine.mean_registrations_per_day() is None

    assert openpyxl.load_workbook(filename=engine.export_excel())


def test_statistics_engine_only_party(stats_env):
    """Tests statistics engine restricted to a set of event types."""
    party = EventType.query.filter_by(name="Soirée").first()
    collective = EventType.query.filter_by(name="Collective").first()

    # Single event type: only the "Soirée" event (event3) is kept.
    engine = StatisticsEngine(event_type_ids=[party.id])
    assert engine.nb_events() == 1
    assert engine.nb_events_by_event_type() == {"Soirée": 1}

    # Set of event types behaves as the union of each type.
    engine = StatisticsEngine(event_type_ids=[party.id, collective.id])
    assert engine.nb_events() == 6
    assert set(engine.nb_events_by_event_type()) == {"Soirée", "Collective"}

    # Empty selection means no restriction on event type.
    engine = StatisticsEngine(event_type_ids=[])
    assert engine.nb_events() == 6
