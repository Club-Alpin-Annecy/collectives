""" " Testing badge management for administration"""

import json
from datetime import date
from io import BytesIO

import openpyxl

from collectives.models import ActivityType, BadgeIds
from tests import utils

# pylint: disable=unused-argument


def test_volunteers_display(supervisor_client, user_with_valid_benevole_badge):
    """Tests display of volunteers for a supervisor user."""
    response = supervisor_client.get("/activity_supervision/volunteers/")
    assert response.status_code == 200

    response = supervisor_client.get("/api/badges/")
    assert response.status_code == 200
    data = json.loads(response.text)
    assert len(data) == 1
    assert data[0]["name"] == BadgeIds.Benevole.display_name()
    assert data[0]["user"]["mail"] == user_with_valid_benevole_badge.mail


def test_volunteers_export(
    supervisor_client, user_with_valid_benevole_badge, user_with_expired_benevole_badge
):
    """Tests the export function for a supervisor."""

    response = supervisor_client.get("/activity_supervision/volunteers/export")
    assert response.status_code == 405

    response = supervisor_client.post("/activity_supervision/volunteers/export")
    assert response.status_code == 302

    response = supervisor_client.get("/activity_supervision/volunteers/")
    data = utils.load_data_from_form(response.text, "badges-export-form")
    data["activity_id"] = ActivityType.query.filter_by(name="Parapente").first().id
    response = supervisor_client.post(
        "/activity_supervision/volunteers/export", data=data
    )
    assert response.status_code == 302

    response = supervisor_client.get("/activity_supervision/volunteers/")
    data = utils.load_data_from_form(response.text, "badges-export-form")
    data["activity_id"] = ActivityType.query.filter_by(name="Alpinisme").first().id
    response = supervisor_client.post(
        "/activity_supervision/volunteers/export", data=data
    )
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 3
    assert worksheet.max_column == 9


def test_volunteers_add(supervisor_client, user3):
    """Tests the ability for a supervisor to add a badge to a user"""
    response = supervisor_client.get("/activity_supervision/volunteers/")
    data = utils.load_data_from_form(response.text, "user-search-form")
    data["user_id"] = str(user3.id)
    data["user_search"] = user3.first_name
    data["activity_id"] = ActivityType.query.filter_by(name="Alpinisme").first().id
    data["badge_id"] = str(int(BadgeIds.Benevole))

    response = supervisor_client.post("/activity_supervision/volunteers/add", data=data)
    assert response.status_code == 302
    assert len(user3.badges) == 1
    assert user3.badges[0].expiration_date > date.today()
    assert user3.badges[0].badge_id == BadgeIds.Benevole


def test_volunteers_bulk_add(supervisor_client, user1, user2, user3):
    """Tests the ability for a supervisor to add a badge to a user"""
    csv = f"license, date, activity\n{user1.license},\n{user2.license},{date.today().isoformat()}\n{user3.license},,Canyon\n"

    response = supervisor_client.get("/activity_supervision/volunteers/")
    data = utils.load_data_from_form(response.text, "user-search-form")
    data["activity_id"] = ActivityType.query.filter_by(name="Alpinisme").first().id
    data["csv_file"] = (BytesIO(csv.encode("utf8")), "import.csv")
    data["badge_id"] = str(int(BadgeIds.Benevole))

    response = supervisor_client.post("/activity_supervision/volunteers/add", data=data)

    assert response.status_code == 302

    assert len(user1.badges) == 1
    assert user1.badges[0].expiration_date > date.today()
    assert user1.badges[0].badge_id == BadgeIds.Benevole
    assert user1.badges[0].activity_name == "Alpinisme"
    assert len(user2.badges) == 1
    assert user2.badges[0].expiration_date > date.today()
    assert user2.badges[0].badge_id == BadgeIds.Benevole
    assert user2.badges[0].activity_name == "Alpinisme"

    # User3 should not have received a badge
    # as supervisor_client is not supervising "Canyon"
    assert len(user3.badges) == 0


def test_volunteers_delete(supervisor_client, user_with_expired_benevole_badge):
    """Tests the ability for a supervisor to delete a user badge."""
    badge_id = user_with_expired_benevole_badge.badges[0].id
    response = supervisor_client.post(
        f"/activity_supervision/volunteers/delete/{badge_id}"
    )
    assert response.status_code == 302
    assert len(user_with_expired_benevole_badge.badges) == 0


def test_volunteers_renew(supervisor_client, user_with_expired_benevole_badge):
    """Tests the ability for a supervisor to renew a user badge."""
    badge_id = user_with_expired_benevole_badge.badges[0].id
    response = supervisor_client.post(
        f"/activity_supervision/volunteers/renew/{badge_id}"
    )
    assert response.status_code == 302
    assert user_with_expired_benevole_badge.badges[0].expiration_date > date.today()
