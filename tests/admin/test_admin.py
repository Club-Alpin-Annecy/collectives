""" " Testing administration"""

from io import BytesIO

import openpyxl
import pytest

from collectives.models import Role, RoleIds, User, db


@pytest.fixture
def user1_with_leader_role(user1):
    """Add EventLeader role to user1 and clean up after test if we added it."""
    from tests.fixtures.user import promote_to_leader

    had_leader_role = user1.has_role([RoleIds.EventLeader])
    if not had_leader_role:
        promote_to_leader(user1)
        db.session.commit()
    yield user1
    if not had_leader_role:
        Role.query.filter_by(user_id=user1.id, role_id=RoleIds.EventLeader).delete()
        db.session.commit()


def load_export_and_check(flask_response, expected_count):
    """Load Excel from response and verify row count and headers."""
    workbook = openpyxl.load_workbook(filename=BytesIO(flask_response.data))
    worksheet = workbook.active
    user_count = worksheet.max_row - 1
    assert user_count == expected_count, (
        f"Expected {expected_count} users, got {user_count}"
    )

    assert worksheet.max_column == 5
    headers = [cell.value for cell in worksheet[1]]
    assert headers == ["Licence", "Prénom", "Nom", "Email", "Téléphone"]

    return worksheet


def test_index(admin_client):
    """Test display of administration index page."""
    response = admin_client.get("/administration/")
    assert response.status_code == 200


def test_user_access(user1_client):
    """Test access refusal to a non admin user."""
    response = user1_client.get("/administration/")
    assert response.status_code == 302


def test_add_user(admin_client):
    """Test display of the user add form."""
    response = admin_client.get("/administration/users/add")
    assert response.status_code == 200


def test_modify_user(admin_client, admin_user):
    """Test display of a user modification form."""
    response = admin_client.get(f"/administration/users/{admin_user.id}")
    assert response.status_code == 200


def test_add_user_role(admin_client, admin_user):
    """Test display of a user role modification form."""
    response = admin_client.get(f"/administration/user/{admin_user.id}/roles")
    assert response.status_code == 200


def test_add_user_badge(admin_client, admin_user):
    """Test display of a user badge modification form."""
    response = admin_client.get(f"/administration/user/{admin_user.id}/badges")
    assert response.status_code == 200


def test_export_roles(admin_client):
    """Test exports of user roles"""
    response = admin_client.get("/administration/roles/export/")
    assert response.status_code == 302

    response = admin_client.get("/administration/roles/export/tnone")
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 2
    assert worksheet.max_column == 7

    response = admin_client.get("/administration/roles/export/t5-r10")
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 1
    assert worksheet.max_column == 7


def test_export_search_results_no_filter(admin_client, user1, user2):
    """
    Test export of all users with no filter.
    user1 and user2 unused parameters ensure that user fixtures are loaded
    """
    expected_count = User.query.count()
    response = admin_client.post("/administration/users/export")
    assert response.status_code == 200
    assert response.content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    load_export_and_check(response, expected_count)


def test_export_search_results_user_access(user1_client):
    """Test access denial to non-admin user."""
    response = user1_client.post("/administration/users/export")
    assert response.status_code == 302


@pytest.mark.parametrize(
    "filter_field,filter_value,expected_count",
    [
        ("mail", "user1@example.org", 1),
        ("mail", "nonexistent@example.org", 0),
        ("first_name", "Jan", 1),
        ("last_name", "Johnston", 1),
    ],
)
def test_export_search_results_with_filter(
    admin_client, user1, user2, filter_field, filter_value, expected_count
):
    """Test export with various filters."""
    response = admin_client.post(
        "/administration/users/export",
        data={"filters[0][field]": filter_field, "filters[0][value]": filter_value},
    )
    assert response.status_code == 200
    load_export_and_check(response, expected_count)


@pytest.mark.parametrize(
    "role_id,expected_count",
    [
        (RoleIds.EventLeader, 1),
        (RoleIds.Administrator, 1),
    ],
)
def test_export_search_results_with_role_filter(
    admin_client, user1_with_leader_role, user2, role_id, expected_count
):
    """Test export with role filters (user1 has EventLeader, admin has Administrator)."""
    filter_value = f"r{int(role_id)}"
    response = admin_client.post(
        "/administration/users/export",
        data={"filters[0][field]": "roles", "filters[0][value]": filter_value},
    )
    assert response.status_code == 200
    load_export_and_check(response, expected_count)


def test_admin_create_valideuser(admin_client):
    """Test  creation a new user."""
    response = admin_client.post(
        "/administration/users/add",
        data={
            "mail": "user1@mail.domain",
            "password": "foobar1+",
            "confirm": "foobar1+",
            "first_name": "user",
            "last_name": "LastName",
            "license": "987987981234",
            "phone": "0123456789",
        },
    )
    assert response.headers["Location"].endswith("/administration/")
    users = User.query.all()
    assert users[1].mail == "user1@mail.domain"


# pylint: disable=unused-argument
def test_create_token(admin_client, extranet_monkeypatch):
    """Test administrator token generation for another user."""
    response = admin_client.get("/administration/token")
    assert response.status_code == 200

    data = {"license": "740020780001", "confirm": "y"}

    response = admin_client.post("/administration/generate_token", data=data)
    assert response.status_code == 200
