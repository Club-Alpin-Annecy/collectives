""" " Testing badge management for administration"""

import json
from datetime import date
from io import BytesIO

import openpyxl

from collectives.forms.activity_type import ActivityTypeSelectionForm
from collectives.models import ActivityType, BadgeIds
from tests import utils

# pylint: disable=unused-argument


def test_admin_badges_display(hotline_client, user_with_valid_benevole_badge):
    """Tests display of badges for an hotline user."""
    response = hotline_client.get("/administration/badges/")
    assert response.status_code == 200

    response = hotline_client.get("/api/badges/")
    assert response.status_code == 200
    data = json.loads(response.text)
    assert len(data) == 1
    assert data[0]["name"] == BadgeIds.Benevole.display_name()
    assert data[0]["user"]["mail"] == user_with_valid_benevole_badge.mail


def test_admin_badges_export(
    hotline_client, user_with_valid_benevole_badge, user_with_expired_benevole_badge
):
    """Tests the export function for an hotline user."""
    response = hotline_client.get("/administration/badges/export/")
    assert response.status_code == 405

    response = hotline_client.post("/administration/badges/export/")
    assert response.status_code == 302

    response = hotline_client.get("/administration/badges/")
    data = utils.load_data_from_form(response.text, "badges-export-form")
    data["activity_id"] = ActivityType.query.filter_by(name="Parapente").first().id

    response = hotline_client.post("/administration/badges/export/", data=data)
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 1
    assert worksheet.max_column == 9

    data["activity_id"] = ActivityTypeSelectionForm.ALL_ACTIVITIES
    response = hotline_client.post("/administration/badges/export/", data=data)
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(filename=BytesIO(response.data))
    worksheet = workbook.active
    assert worksheet.max_row == 3


def test_admin_badges_add(hotline_client, user3):
    """Tests the ability for an hotline user to add a badge to a user"""
    response = hotline_client.get("/administration/badges/")
    data = utils.load_data_from_form(response.text, "user-search-form")
    data["user_id"] = str(user3.id)
    data["user_search"] = user3.first_name
    data["activity_id"] = ActivityType.query.filter_by(name="Alpinisme").first().id
    data["badge_id"] = str(int(BadgeIds.Benevole))

    response = hotline_client.post("/administration/badges/add", data=data)
    assert response.status_code == 302
    assert len(user3.badges) == 1
    assert user3.badges[0].expiration_date > date.today()
    assert user3.badges[0].badge_id == BadgeIds.Benevole


def test_admin_badges_delete(hotline_client, user_with_expired_benevole_badge):
    """Tests the ability for an hotline user to delete a user badge."""
    badge_id = user_with_expired_benevole_badge.badges[0].id
    response = hotline_client.post(f"/administration/badges/delete/{badge_id}")
    assert response.status_code == 302
    assert len(user_with_expired_benevole_badge.badges) == 0


def test_admin_badges_renew(hotline_client, user_with_expired_benevole_badge):
    """Tests the abilty for an hotline user to renew the badge for a year"""
    badge_id = user_with_expired_benevole_badge.badges[0].id
    response = hotline_client.post(f"/administration/badges/renew/{badge_id}")
    assert response.status_code == 302
    assert user_with_expired_benevole_badge.badges[0].expiration_date > date.today()
